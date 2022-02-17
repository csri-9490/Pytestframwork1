from selenium import webdriver

import openpyxl
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
# from selenium.webdriver.common.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
book=openpyxl.load_workbook("C:\\Users\\srika\\OneDrive\\Desktop\\details.xlsx")
# sheet=book.active
sheet=book['Sheet_Name']
print(sheet)
col=sheet.max_column
row=sheet.max_row
print(row)
print(col)
c1=[]
c2=[]
driver=webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.get("https://rahulshettyacademy.com/angularpractice/")
for i in range(2,row+1):
     for j in range(1,col):
        c1.append(sheet.cell(row=i,column=j).value)

for k in range(2,row+1):
    for l in range(2,col+1):
        # c2.append(sheet.cell(row=k,column=l).value)
        a1=sheet.cell(row=k,column=l).value
        driver.find_element(By.XPATH, "//input[@name='email']").send_keys(a1)
        time.sleep(5)
        driver.refresh()
        time.sleep(5)

print('a1 value is',a1)
for p in range(2,row+1):
    a11=sheet.cell(row=p,column=1).value
# print(c1)
# print(c2)
print('a11 is',a11)
# driver=webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# driver.get("https://rahulshettyacademy.com/angularpractice/")

time.sleep(5)
# driver.find_element(By.CSS_SELECTOR,"input[class='form-control ng-pristine ng-invalid ng-touched']").send_keys(c1)
# driver.find_element(By.XPATH,"//input[@name='email']").clear()
# driver.find_element(By.XPATH,"//input[@name='email']").send_keys(a1)
time.sleep(5)
# driver.refresh()
time.sleep(5)