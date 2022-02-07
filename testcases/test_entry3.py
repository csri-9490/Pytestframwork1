import openpyxl

# file = open("C:\\Users\\srika\\OneDrive\\Desktop\\details.xls","w+")
from openpyxl.styles import Font, colors

book=openpyxl.load_workbook("C:\\Users\\srika\\OneDrive\\Desktop\\details.xlsx")
sheet=book.active
# sheet=book.create_sheet('second')
print(sheet)
print(sheet.active_cell)
print((sheet.cell))
cell=sheet['A5']
print(sheet['A5'].value)
cell=sheet.cell(row=4,column=2)
print(cell.value)
print(book.sheetnames)
# print(book.sheetnames[0])

